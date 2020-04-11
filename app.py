import time
from flask import Flask, render_template, request,send_from_directory,session
import pandas as pd
import re
from itertools import chain
import netaddr
from netaddr import *
import os

from openpyxl import load_workbook

app = Flask(__name__)

basedir = os.path.abspath(os.path.dirname(__file__))
print(basedir)

path1 = os.path.join(basedir,'Upload_folder\\')
path2 = os.path.join(basedir,'TDD_folder\\')
print(path1)
print(path2)
app.config['UPLOAD_FOLDER'] = path1
app.config['TDD_FOLDER'] = path2
print(app.config['UPLOAD_FOLDER'])
print(app.config['TDD_FOLDER'])

if os.path.isdir(path1):
    print('ok')
else:
    print("not ok Path1")
    os.mkdir(path1)

if os.path.isdir(path2):
    print('ok')
else:
    print("not ok Path2")
    os.mkdir(path2)

app.secret_key='vil123'



def get_timestamp():
    return int(time.time())
app.jinja_env.globals['timestamp'] = get_timestamp

@app.route('/')
def index():
    return render_template('home.html')

@app.route('/check')
def check():
    return render_template('test.html')

@app.route('/result',methods = ['POST'])
def result():
    if request.method == 'POST':
        f1 = request.files['file1']
        f1.save(os.path.join(app.config['UPLOAD_FOLDER'],f1.filename))
        f2 = request.files['file2']
        f2.save(os.path.join(app.config['UPLOAD_FOLDER'],f2.filename))
        f3 = request.files['file3']
        f3.save(os.path.join(app.config['UPLOAD_FOLDER'],f3.filename))
        f4 = request.files['file4']
        f4.save(os.path.join(app.config['UPLOAD_FOLDER'],f4.filename))
        coreopt=request.form['optradio']
        print(coreopt)
        session['f1']=  f1.filename
        session['f2'] = f2.filename
        session['f3'] = f3.filename
        session['f4'] = f4.filename
        my_func(f1,f2,f3,f4,coreopt)
        f1.close()
        f2.close()
        f3.close()
        f4.close()
        return render_template('file-downloads.html',name1=f4.filename,name2=f2.filename)

@app.route('/file-downloads/<name1>')
def file_downloads(name1):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename=name1, as_attachment=True)

@app.route('/delete/')
def delete():
    os.remove(os.path.join(app.config['UPLOAD_FOLDER'], session.get('f1',None)))
    os.remove(os.path.join(app.config['UPLOAD_FOLDER'], session.get('f2',None)))
    os.remove(os.path.join(app.config['UPLOAD_FOLDER'], session.get('f3', None)))
    os.remove(os.path.join(app.config['UPLOAD_FOLDER'], session.get('f4', None)))
    print("All Files Deleted")
    return render_template('delete.html')

@app.route('/tdd')
def tdd():
    return render_template('tdd.html')

@app.route('/tddpost',methods = ['POST'])
def tddpost():
    if request.method == 'POST':
        tddplan = request.files['tddfile1']
        tddplan.save(os.path.join(app.config['TDD_FOLDER'],tddplan.filename))
        tddtemplate = request.files['tddfile2']
        tddtemplate.save(os.path.join(app.config['TDD_FOLDER'],tddtemplate.filename))
        tddfinalfile = request.files['tddfile3']
        tddfinalfile.save(os.path.join(app.config['TDD_FOLDER'],tddfinalfile.filename))
        tddopt=request.form['optradio']
        print(tddopt)
        session['tddplan'] =  tddplan.filename
        session['tddtemplate'] = tddtemplate.filename
        session['tddfinalfile'] = tddfinalfile.filename
        print(tddplan.filename)
        print(tddtemplate.filename)
        print(tddfinalfile.filename)
        if tddopt=='PRE-AGG':
            TDD(tddplan, tddtemplate, tddfinalfile)
        else:
            TDD_IPRAN(tddplan,tddtemplate,tddfinalfile)
        tddplan.close()
        tddtemplate.close()
        tddfinalfile.close()

        return render_template('file-downloads_tdd.html',name1=tddfinalfile.filename)

@app.route('/deletetdd/')
def deletetdd():
    os.remove(os.path.join(app.config['TDD_FOLDER'], session.get('tddplan',None)))
    os.remove(os.path.join(app.config['TDD_FOLDER'], session.get('tddtemplate',None)))
    os.remove(os.path.join(app.config['TDD_FOLDER'], session.get('tddfinalfile', None)))
    print("All TDD Files Deleted")
    return render_template('delete.html')

@app.route('/file-downloads_tdd/<name1>')
def file_downloads_tdd(name1):
    return send_from_directory(app.config['TDD_FOLDER'], filename=name1, as_attachment=True)

@app.route('/ciscoipran')
def ciscoipran():
    return render_template('ciscoipran.html')

@app.route('/ciscopreagg')
def ciscopreagg():
    return render_template('ciscopreagg.html')

@app.route('/juniperagg1')
def juniperagg1():
    return render_template('juniperagg1.html')

@app.route('/juniperagg2')
def juniperagg2():
    return render_template('juniperagg2.html')

def my_func(a1,a2,a3,a4,coreopt):

    DS = pd.read_excel(a1, sheet_name='Data Sheet')
    from openpyxl import load_workbook
    logfile2 = list(open(os.path.join(app.config['UPLOAD_FOLDER'],a2.filename), 'r').read().split('\n'))
    logfile1 = list(open(os.path.join(app.config['UPLOAD_FOLDER'],a3.filename), 'r').read().split('\n'))
    CORE = coreopt  # IDEA or VODA
    print('@@@@',CORE)
    idea = DS[DS['Router Owner'] == 'E-IDEA']
    voda = DS[DS['Router Owner'] == 'E-VF']
    idr_vbsc = idea[idea['BSC'] == 'VODA']
    idr_vrnc = idea[idea['RNC'] == 'VODA']
    vor_ibsc = voda[voda['BSC'] == 'IDEA']
    vor_irnc = voda[voda['RNC'] == 'IDEA']
    live_site = idea[idea['Site Type'] == 'SRAN Live']
    df2 = pd.DataFrame()

    def check(S, S_OUT):
        l1 = []
        l2 = []
        l3 = []
        l4 = []
        l5 = []
        l6 = []

        for i in S_OUT:
            if (i.prefixlen < 28):
                l1.append(i)
            else:
                l2.append(i)

        for j in l1:
            l3.append(list(j.subnet(30)))
            l4 = list(chain.from_iterable(l3))

        for k in S:
            if (k not in l2):
                l5.append(k)
        for m in l5:
            if m not in l4:
                l6.append(m)
        return l6

    vf_pabis_out = 'mob_Abis_IDEA_NNI_HB_OUT'
    vf_pabis_in = 'mob_Abis_IDEA_NNI_HB_IN'
    vf_iub_out = 'mob_iub_IDEA_NNI_HB_OUT'
    vf_iub_in = 'mob_iub_IDEA_NNI_HB_IN'
    vf_oam_in = 'HUAWEI_TDD_ENB_OAM_IN'
    vf_x2_out = 'MOB_ENB_X2_IDEA_NNI_HB_OUT'
    vf_x2_in = 'MOB_ENB_X2_IDEA_NNI_HB_IN'
    vf_s1u_out = 'eNB_SGW_S1U_IDEA_HB_OUT'
    vf_s1u_in = 'eNB_SGW_S1U_IDEA_HB_IN'
    vf_s1c_out = 'IDEA_IPRAN_ENB_S1C_OUT'
    vf_s1c_in = 'IDEA_IPRAN_ENB_S1C_IN'

    def c2l(logfile2, str):
        flag = False
        summary = ''
        l = []
        for line in logfile2:
            if str in line:
                flag = True
            elif flag:
                summary += line
                if not line.strip(): break

        ips = re.findall(r'\d+.\d+.\d+.\d+/\d+', summary)
        for ip in ips:
            l.append(ip)
        return l

    VF_PABIS_OUT = c2l(logfile2, vf_pabis_out)
    VF_IUB_OUT = c2l(logfile2, vf_iub_out)
    VF_OAM_IN = c2l(logfile2, vf_oam_in)
    VF_S1C_OUT = c2l(logfile2, vf_s1c_out)
    VF_S1U_OUT = c2l(logfile2, vf_s1u_out)
    VF_X2_OUT = c2l(logfile2, vf_x2_out)

    VF_PABIS_IN = c2l(logfile2, vf_pabis_in)
    VF_IUB_IN = c2l(logfile2, vf_iub_in)
    VF_S1C_IN = c2l(logfile2, vf_s1c_in)
    VF_S1U_IN = c2l(logfile2, vf_s1u_in)
    VF_X2_IN = c2l(logfile2, vf_x2_in)

    id_pabis_out = 'prefix-set VODA_PABIS_OUT'
    id_pabis_in = 'prefix-set VODA_PABIS_IN'
    id_iub_out = 'prefix-set VODA_IuB_OUT'
    id_iub_in = 'prefix-set VODA_IuB_IN'
    id_oam_out = 'prefix-set VODA_TDD_OUT'
    id_x2_out = 'prefix-set VODA_X2_OUT'
    id_x2_in = 'prefix-set VODA_X2_IN'
    id_s1u_out = 'prefix-set IDEA_SGW_OUT'
    id_s1u_in = 'prefix-set VF_SGW_IN'
    id_s1c_out = 'prefix-set VODA_S1-MME_OUT'
    id_s1c_in = 'prefix-set VODA_S1-MME_IN'

    def c2l(logfile1, str):
        flag = False
        summary = ''
        l = []
        for line in logfile1:
            if line.startswith(str):
                flag = True
            elif flag:
                summary += line
                if not line.strip(): break

        ips = re.findall(r'\d+.\d+.\d+.\d+/\d+', summary)
        for ip in ips:
            l.append(ip)
        return l

    ID_PABIS_OUT = c2l(logfile1, id_pabis_out)
    ID_IUB_OUT = c2l(logfile1, id_iub_out)
    ID_OAM_OUT = c2l(logfile1, id_oam_out)
    ID_S1C_OUT = c2l(logfile1, id_s1c_out)
    ID_S1U_OUT = c2l(logfile1, id_s1u_out)
    ID_X2_OUT = c2l(logfile1, id_x2_out)

    ID_PABIS_IN = c2l(logfile1, id_pabis_in)
    ID_IUB_IN = c2l(logfile1, id_iub_in)
    ID_S1C_IN = c2l(logfile1, id_s1c_in)
    ID_S1U_IN = c2l(logfile1, id_s1u_in)
    ID_X2_IN = c2l(logfile1, id_x2_in)

    def write(dfname, sheetid, colf):
        print(dfname)

        wb = load_workbook(os.path.join(app.config['UPLOAD_FOLDER'],a4.filename))

        ws = wb[sheetid]
        print(ws)
        for index, row in dfname.iterrows():
            cell = colf + '%d' % (index + 2)
            ws[cell] = row[0]
        dfname.drop(dfname.index, inplace=True)
        wb.save(os.path.join(app.config['UPLOAD_FOLDER'],a4.filename))

    def IDEAFO(str_name, rl, service, io, op, col):
        plan = []
        WO = []
        x = []
        for i in idea[str_name]:
            plan.append(IPNetwork(i))

        for i in rl:
            WO.append(IPNetwork(i))

        i_w_out = check(plan, WO)
        print('{} - {}_{}_NETWORKS'.format(op, service, io))
        for i in i_w_out:
            x.append(str(netaddr.IPNetwork(i)))
        for i in range(len(x)):
            df2.loc[i, 0] = x[i]
        print(len(x))
        write(df2, op, col)

    def VODAFO(str_name, rl, service, io, op, col):
        plan = []
        WO = []
        x = []
        for i in voda[str_name]:
            plan.append(IPNetwork(i))

        for i in rl:
            WO.append(IPNetwork(i))

        i_w_out = check(plan, WO)
        print('{} - {}_{}_NETWORKS'.format(op, service, io))
        for i in i_w_out:
            x.append(str(netaddr.IPNetwork(i)))
        for i in range(len(x)):
            df2.loc[i, 0] = x[i]
        print(len(x))
        write(df2, op, col)

    def VORIBSC(str_name, rl, service, io, op, col1):
        plan = []
        WO = []
        x = []
        for i in vor_ibsc[str_name]:
            plan.append(IPNetwork(i))

        for i in rl:
            WO.append(IPNetwork(i))

        i_w_out = check(plan, WO)
        print('{} - {}_{}_NETWORKS'.format(op, service, io))
        for i in i_w_out:
            x.append(str(netaddr.IPNetwork(i)))
        for i in range(len(x)):
            df2.loc[i, 0] = x[i]
        print(len(x))
        write(df2, op, col1)

    def IDRVBSC(str_name, rl, service, io, op, col):
        plan = []
        WO = []
        x = []
        for i in idr_vbsc[str_name]:
            plan.append(IPNetwork(i))

        for i in rl:
            WO.append(IPNetwork(i))

        i_w_out = check(plan, WO)
        print('{} - {}_{}_NETWORKS'.format(op, service, io))
        for i in i_w_out:
            x.append(str(netaddr.IPNetwork(i)))
        for i in range(len(x)):
            df2.loc[i, 0] = x[i]
        print(len(x))
        write(df2, op, col)

    def IDRVRNC(str_name, rl, service, io, op, col):
        plan = []
        WO = []
        x = []
        for i in idr_vrnc[str_name]:
            plan.append(IPNetwork(i))

        for i in rl:
            WO.append(IPNetwork(i))

        i_w_out = check(plan, WO)
        print('{} - {}_{}_NETWORKS'.format(op, service, io))
        for i in i_w_out:
            x.append(str(netaddr.IPNetwork(i)))
        for i in range(len(x)):
            df2.loc[i, 0] = x[i]
        print(len(x))
        write(df2, op, col)

    def VORIRNC(str_name, rl, service, io, op, col):
        plan = []
        WO = []
        x = []
        for i in vor_irnc[str_name]:
            plan.append(IPNetwork(i))

        for i in rl:
            WO.append(IPNetwork(i))

        i_w_out = check(plan, WO)
        print('{} - {}_{}_NETWORKS'.format(op, service, io))
        for i in i_w_out:
            x.append(str(netaddr.IPNetwork(i)))
        for i in range(len(x)):
            df2.loc[i, 0] = x[i]
        print(len(x))
        write(df2, op, col)

    def SRANLIVE(str_name, rl, service, io, op, col2):
        plan = []
        WO = []
        x = []
        for i in live_site[str_name]:
            plan.append(IPNetwork(i))

        for i in rl:
            WO.append(IPNetwork(i))

        i_w_out = check(plan, WO)
        print('{} - {}_{}_NETWORKS'.format(op, service, io))
        for i in i_w_out:
            x.append(str(netaddr.IPNetwork(i)))
        for i in range(len(x)):
            df2.loc[i, 0] = x[i]
        print(len(x))
        write(df2, op, col2)

    # IUB networks Idea router-----Voda RNC

    if (idr_vrnc.empty == True):
        print('nothing to add in CISCO IUB')
    else:
        IDRVRNC('3G-NodeB Network', ID_IUB_OUT, 'IUB_NETWORKS', 'OUT', 'IDEA', 'D')
        IDRVRNC('3G-NodeB Network', VF_IUB_IN, 'IUB_NETWORKS', 'IN', 'VODA', 'C')

    # IUB networks Voda router-----Idea RNC

    if (vor_irnc.empty == True):
        print('nothing to add in JUNIPER IUB')
    else:
        VORIRNC('3G-NodeB Network', VF_IUB_OUT, 'IUB_NETWORKS', 'OUT', 'VODA', 'C')
        VORIRNC('3G-NodeB Network', ID_IUB_IN, 'IUB_NETWORKS', 'IN', 'IDEA', 'D')

    # PABIS networks idea router-----Voda BSC

    if (idr_vbsc.empty == True):
        print('nothing to add in CISCO PABIS')
    else:
        IDRVBSC('2G-BTS Network', ID_PABIS_OUT, 'PABIS_NETWORKS', 'OUT', 'IDEA', 'B')
        IDRVBSC('2G-BTS Network', VF_PABIS_IN, 'PABIS_NETWORKS', 'IN', 'VODA', 'A')

    # PABIS networks Voda router---idea BSC

    if (vor_ibsc.empty == True):
        print('nothing to add in JUNIPER PABIS')
    else:
        VORIBSC('2G-BTS Network', VF_PABIS_OUT, 'PABIS_NETWORKS', 'OUT', 'VODA', 'B')
        VORIBSC('2G-BTS Network', ID_PABIS_IN, 'PABIS_NETWORKS', 'IN', 'IDEA', 'A')

    # OAM

    IDEAFO('OAM Network', ID_OAM_OUT, 'OAM_NETWORKS', 'OUT', 'IDEA', 'F')
    IDEAFO('OAM Network', VF_OAM_IN, 'OAM_NETWORKS', 'IN', 'VODA', 'E')

    # S1-MME

    if CORE == 'IDEA':
        print('###', CORE)
        VODAFO('S1-C Network', VF_S1C_OUT, 'S1-MME_NETWORKS', 'OUT', 'VODA', 'H')
        VODAFO('S1-C Network', ID_S1C_IN, 'S1-MME_NETWORKS', 'IN', 'IDEA', 'G')
    else:
        IDEAFO('S1-C Network', ID_S1C_OUT, 'S1-MME_NETWORKS', 'OUT', 'IDEA', 'H')
        IDEAFO('S1-C Network', VF_S1C_IN, 'S1-MME_NETWORKS', 'IN', 'VODA', 'G')

    # IDEA ROUTER S1-U

    IDEAFO('S1-U Network', ID_S1U_OUT, 'S1-U_NETWORKS', 'OUT', 'IDEA', 'J')  # I-S1-U_OUT
    IDEAFO('S1-U Network', VF_S1U_IN, 'S1-U_NETWORKS', 'IN', 'VODA', 'I')  # V-S1-U_IN

    # VODA ROUTER S1-U

    VODAFO('S1-U Network', VF_S1U_OUT, 'S1-U_NETWORKS', 'OUT', 'VODA', 'J')  # V-S1-U_OUT
    VODAFO('S1-U Network', ID_S1U_IN, 'S1-U_NETWORKS', 'IN', 'IDEA', 'I')  # I-S1-U_IN

    # IDEA ROUTER X2

    IDEAFO('X2 Network', ID_X2_OUT, 'X2_NETWORKS', 'OUT', 'IDEA', 'L')  # I-X2_OUT
    IDEAFO('X2 Network', VF_X2_IN, 'X2_NETWORKS', 'IN', 'VODA', 'K')  # V-X2_IN

    # VODA ROUTER X2

    VODAFO('X2 Network', VF_X2_OUT, 'X2_NETWORKS', 'OUT', 'VODA', 'L')  # V-X2_OUT
    VODAFO('X2 Network', ID_X2_IN, 'X2_NETWORKS', 'IN', 'IDEA', 'K')  # I-X2_IN

    # SRAN LIVE SITE

    if CORE == 'IDEA':

        if (live_site['S1-C Network'].empty == True):
            print('NOTHING TO ADD IN CISCO')

        else:
            SRANLIVE('S1-C Network', ID_S1C_OUT, 'SRAN-LIVE SITE----S1-MME_NETWORKS', 'OUT', 'IDEA', 'N')
            SRANLIVE('S1-C Network', VF_S1C_IN, 'SRAN-LIVE SITE----S1-MME_NETWORKS', 'IN', 'VODA', 'M')


def TDD(tddplanf,tddtemplatef,tddfinalfilef):
    df = pd.read_excel(tddplanf)
    print(df.head())

    logfile = open(os.path.join(app.config['TDD_FOLDER'], tddtemplatef.filename), 'r')
    lines = logfile.readlines()

    df2 = pd.DataFrame(columns=['Router', 'Script'])
    f = df['TDD-LTE VLANs'].count()
    s = []

    for i in range(0, f):

        for line in lines:
            line = re.sub('DIST', df.at[i, 'District'], line.rstrip())
            line = re.sub('POP', df.at[i, 'OF PoP Name'], line.rstrip())
            line = re.sub('ROUTER', df.at[i, 'Router Location Name'], line.rstrip())
            line = re.sub('RFID', df.at[i, 'Infra ID'], line.rstrip())
            line = re.sub('SNAME', df.at[i, 'Site Name'], line.rstrip())
            line = re.sub('vlan1', str(df.at[i, 'VLAN1']), line.rstrip())
            line = re.sub('vlan2', str(df.at[i, 'VLAN2']), line.rstrip())
            line = re.sub('vlan3', str(df.at[i, 'VLAN3']), line.rstrip())
            line = re.sub('ip1', df.at[i, 'Gateway IP1'], line.rstrip())
            line = re.sub('ip2', df.at[i, 'Gateway IP2'], line.rstrip())
            line = re.sub('ip3', df.at[i, 'Gateway IP3'], line.rstrip())
            line = re.sub('MASK', '255.255.255.252', line.rstrip())
            line = re.sub('PORTID', str(df.at[i, 'Router Port']), line.rstrip())
            s.append(line)
        for k in range(0, len(s)):
            df2.loc[k, 'Router'] = df.at[i, 'Router Location Name']
            df2.loc[k, 'Script'] = s[k]
        s = []
        book = load_workbook(os.path.join(app.config['TDD_FOLDER'], tddfinalfilef.filename))
        writer = pd.ExcelWriter(os.path.join(app.config['TDD_FOLDER'], tddfinalfilef.filename), engine='openpyxl')
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}

        for sheetname in writer.sheets:
            df2.to_excel(writer, sheet_name='Sheet1', startrow=writer.sheets['Sheet1'].max_row, index=False,
                         header=False)
        print(df2)
        df2.drop(df2.index, inplace=True)
        writer.save()


def TDD_IPRAN(tddplanf,tddtemplatef,tddfinalfilef):
    df = pd.read_excel(tddplanf)

    logfile = open(os.path.join(app.config['TDD_FOLDER'], tddtemplatef.filename), 'r')
    lines = logfile.readlines()

    df2 = pd.DataFrame(columns=['Router', 'Script'])
    f = df['TDD-LTE VLANs'].count()
    s = []

    for i in range(0, f):

        for line in lines:
            line = re.sub('DIST', df.at[i, 'District'], line.rstrip())
            line = re.sub('POP', df.at[i, 'OF PoP Name'], line.rstrip())
            line = re.sub('ROUTER', df.at[i, 'Router Location Name'], line.rstrip())
            line = re.sub('RFID', df.at[i, 'Infra ID'], line.rstrip())
            line = re.sub('SNAME', df.at[i, 'Site Name'], line.rstrip())
            line = re.sub('vlan1', str(df.at[i, 'VLAN1']), line.rstrip())
            line = re.sub('vlan2', str(df.at[i, 'VLAN2']), line.rstrip())
            line = re.sub('vlan3', str(df.at[i, 'VLAN3']), line.rstrip())
            line = re.sub('ip1', df.at[i, 'Gateway IP1'], line.rstrip())
            line = re.sub('ip2', df.at[i, 'Gateway IP2'], line.rstrip())
            line = re.sub('ip3', df.at[i, 'Gateway IP3'], line.rstrip())
            line = re.sub('MASK', '255.255.255.252', line.rstrip())
            line = re.sub('BUNDLENAME', str(df.at[i, 'Router Port']), line.rstrip())
            s.append(line)
        for k in range(0, len(s)):
            df2.loc[k, 'Router'] = df.at[i, 'Router Location Name']
            df2.loc[k, 'Script'] = s[k]
        s = []
        book = load_workbook(os.path.join(app.config['TDD_FOLDER'], tddfinalfilef.filename))
        writer = pd.ExcelWriter(os.path.join(app.config['TDD_FOLDER'], tddfinalfilef.filename), engine='openpyxl')
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}

        for sheetname in writer.sheets:
            df2.to_excel(writer, sheet_name='Sheet1', startrow=writer.sheets['Sheet1'].max_row, index=False,
                         header=False)
        print(df2)
        df2.drop(df2.index, inplace=True)
        writer.save()

if __name__ == '__main__':
    app.run(debug = False)
